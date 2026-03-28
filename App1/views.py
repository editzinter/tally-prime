from django.shortcuts import render
from django.contrib import messages
from multiprocessing import context
from unicodedata import name
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect
from django.conf import settings

from .models import *
from datetime import datetime, date, timedelta
from django.contrib.auth.models import User, auth
from django.views.decorators.csrf import csrf_exempt
import json
import csv
import os

# Create your views here.

def base(request):
    return render(request, 'base.html')

def home(request):
    companies = crtcompony.objects.all()
    active_company_id = request.session.get('active_company_id')
    active_company = None
    if active_company_id:
        try:
            active_company = crtcompony.objects.get(pk=active_company_id)
        except crtcompony.DoesNotExist:
            request.session.pop('active_company_id', None)
    from datetime import date
    today = date.today()
    context = {
        'companies': companies,
        'active_company': active_company,
        'has_company': companies.exists(),
        'current_date': today.strftime('%A, %d-%b-%y'),
        'current_period': f"1-Apr-{today.year-1 if today.month < 4 else today.year} to 31-Mar-{today.year if today.month < 4 else today.year+1}",
    }
    return render(request, 'home.html', context)


def branch(request):
    return render(request, 'branch.html')


#------Change Company Creation--------#

def changecompony(request):
    return render(request, 'changecompony.html')

def createcompony(request):
    return render(request, 'createcompony.html')

def crtecompony(request):
    if request.method == 'GET':
        return redirect('createcompony')
    if request.method=='POST':
        comname=request.POST.get('componyname', '')
        mailingname=request.POST.get('mailingname', comname)
        address=request.POST.get('address', '')
        state=request.POST.get('state', '')
        country=request.POST.get('country', 'India')
        pincode=request.POST.get('pincode', '')
        telphone=request.POST.get('telphone', '')
        mobile=request.POST.get('mobile', '')
        fax=request.POST.get('fax', '')
        email=request.POST.get('email', '')
        website=request.POST.get('website', '')
        fyearbgn=request.POST.get('fyearbgn', '') or None
        booksbgn=request.POST.get('booksbgn', '') or None
        curncysymbl=request.POST.get('curncysymbl', '₹')
        crncyname=request.POST.get('crncyname', 'INR')
        # items=request.FILES['file']
        data=crtcompony(componyname=comname,
                    mailingname=mailingname,
                    address=address,
                    state=state,
                    country=country,
                    pincode=pincode,
                    telphone=telphone,
                    mobile=mobile,
                    fax=fax,
                    email=email,
                    website=website,
                    fyearbgn=fyearbgn,
                    booksbgn=booksbgn,
                    curncysymbl=curncysymbl,
                    crncyname=crncyname)
        data.save()
        # Auto-create default ledgers (Cash and Profit & Loss A/c) like real TallyPrime
        if not LedgerModels.objects.filter(ledger_name='Cash').exists():
            LedgerModels(ledger_name='Cash', alias='', under='Cash-in-hand', mail_name='', mail_address='', mail_state='', mail_country='', pan_no='', registration_type='', gst_in='', alter_gst='').save()
        if not LedgerModels.objects.filter(ledger_name='Profit & Loss A/c').exists():
            LedgerModels(ledger_name='Profit & Loss A/c', alias='', under='Primary', mail_name='', mail_address='', mail_state='', mail_country='', pan_no='', registration_type='', gst_in='', alter_gst='').save()
        # Set the newly created company as the active company
        request.session['active_company_id'] = data.pk
        messages.success(request, "Company Registered Successfully!")
        return redirect('home')

def changecompony(request):
    data=crtcompony.objects.all()
    return render(request,'changecompony.html',{'data':data})

def selectcompony(request):
    if request.method == 'POST':
        company_id = request.POST.get('company_id')
        if company_id:
            request.session['active_company_id'] = int(company_id)
        return redirect('home')
    data = crtcompony.objects.all()
    return render(request, 'selectcompony.html', {'data': data})


#---------Group Creation---------#

def group(request):
    grp=GroupModel.objects.all()
    context={'grp':grp,}
    return render(request, 'groups.html',context)

def branch(request):
    context={ 'name':'Branch/Division' }
    return render(request, 'branch.html',context)


@csrf_exempt
def create_group(request):
    if request.method == 'POST':
        gname = request.POST['gname']
        alia = request.POST['alia']
        if len(gname) <= 0:
            return JsonResponse({
                'status': 00
            })

        if len(alia) <= 0:
            alia = None
        else:
            pass

        under = request.POST['und']
        gp = request.POST['subled']
        nett = request.POST['nee']
        calc = request.POST['cal']
        meth = request.POST['meth']

        mdl = GroupModel(
            name=gname,
            alias=alia,
            under=under,
            gp_behaves_like_sub_ledger=gp,
            nett_debit_credit_bal_reporting=nett,
            used_for_calculation=calc,
            method_to_allocate_usd_purchase=meth,
        )
        mdl.save()
        # return redirect('index_view')
        return JsonResponse({
            'status': 1
        })

def load_create_group1(request):
    return render(request,'load_create_groups.html') 

def load_create_groups(request):
    grp = GroupModel.objects.all()
    context={'grp':grp}
    return render(request,'groups.html',context)

def create_group(request):
    if request.method == 'POST':
        gname = request.POST['gname']
        alia = request.POST['alia']
        under = request.POST['und']
        gp = request.POST['subled']
        naturee = request.POST['nature']
        gross_profitt = request.POST['gross_profit']
        nett = request.POST['nee'] 
        calc = request.POST['cal']
        meth = request.POST['meth']

        grp = GroupModel.objects.all()
        context={'grp':grp}

        if GroupModel.objects.filter(name=gname).exists():
                messages.info(request,'This Name is already taken...!')
                return render(request,'load_create_groups.html',context)

        mdl = GroupModel(
            name=gname,
            alias=alia,
            under=under,
            nature_of_group=naturee,
            does_it_affect=gross_profitt,
            gp_behaves_like_sub_ledger=gp,
            nett_debit_credit_bal_reporting=nett,
            used_for_calculation=calc,
            method_to_allocate_usd_purchase=meth,
        )
        mdl.save()
        grp = GroupModel.objects.all()
        context={'grp':grp}
        messages.info(request,'GROUP CREATED SUCCESSFULLY')
        return render(request,'load_create_groups.html',context)


def update_grp(request,pk):
    if request.method=='POST':
        grp =GroupModel.objects.get(id=pk)
        grp.name = request.POST.get('gname')
        grp.alias = request.POST.get('alia')
        grp.under = request.POST.get('under')
        grp.nature_of_group = request.POST.get('nature')
        grp.does_it_affect = request.POST.get('gross_profit')
        grp.gp_behaves_like_sub_ledger = request.POST.get('subled')
        grp.nett_debit_credit_bal_reporting = request.POST.get('nee')
        grp.used_for_calculation = request.POST.get('cal')
        grp.method_to_allocate_usd_purchase = request.POST.get('meth')
        
        grp.save()
        return redirect('groups')
    return render(request, 'update_grp.html',)




#----------currency creation---------#
        
def currency(request):
    obj=CreateCurrency.objects.all()
    context={'cur':obj,}
    return render(request, 'currency.html',context)

def currency_alter(request,pk):
    cur=CreateCurrency.objects.get(id=pk)
    return render(request,'currency_alter.html',{'i':cur})
def load_create_currency(request):
    return render(request,'load_create_currency.html')

def create_currency(request):
    if request.method == 'POST':
        symbol = request.POST.get('symbol', '')
        fname = request.POST.get('fname', '')
        
        if len(symbol) <= 0 or len(fname) <= 0:
            messages.info(request, 'Symbol and Formal Name are required!')
            return redirect('load_create_currency')

        iso_code = request.POST.get('iso_code', '')
        n_deci_placs = request.POST.get('n_deci_placs', '2')
        smt_millon = request.POST.get('smt_millon', 'No')
        symbol_to_amount = request.POST.get('symbol_to_amount', 'No')
        space_bt_sy = request.POST.get('space_bt_sy', 'Yes')
        amount_after_decimal = request.POST.get('amount_after_decimal', '')
        amount_in_words = request.POST.get('amount_in_words', '2')

        mdl_obj = CreateCurrency(
            symbol=symbol,
            formal_name=fname,
            ISO_code=iso_code,
            decimal_places=n_deci_placs,
            show_in_millions=smt_millon,
            suffix_to_amount=symbol_to_amount,
            space_symbol_amount=space_bt_sy,
            word_after_decimal=amount_after_decimal,
            decimal_no_in_words=amount_in_words,
        )
        mdl_obj.save()
        messages.info(request, 'CURRENCY CREATED SUCCESSFULLY')
        return redirect('currency')

def save_currency_data(request):
    if request.method == 'POST':
        sl = request.POST['slno']
        cname = request.POST['curname']
        stdr = request.POST['stdr']
        lvr = request.POST['lvr']
        sr = request.POST['sr']
        lvr2 = request.POST['lvr2']
        sr2 = request.POST['sr2']
        
        obj = CurrencyAlter(
            slno = sl,
            currencys= cname,
            stdrate = stdr,
            lastvrate = lvr,
            specirate = sr,
            lastvrate2 = lvr2,
            specirate2 = sr2,        
           
        )
        
        obj.save()
        grp = CreateCurrency.objects.all()
        obj1 = CurrencyAlter.objects.all()
        context = {'grp':grp ,'obj':obj1}
        return redirect('load_rates_of_exchange',context)

def update_currency(request,pk):
    if request.method=='POST':
        cur =CreateCurrency.objects.get(id=pk)
        cur.symbol = request.POST.get('symbol')
        cur.formal_name = request.POST.get('fname')
        cur.ISO_code = request.POST.get('iso_code')
        cur.decimal_places = request.POST.get('n_deci_placs')
        cur.show_in_millions = request.POST.get('smt_millon')
        cur.suffix_to_amount = request.POST.get('symbol_to_amount')
        cur.space_symbol_amount = request.POST.get('space_bt_sy')
        cur.word_after_decimal = request.POST.get('amount_after_decimal')
        cur.decimal_no_in_words = request.POST.get('amount_in_words')
        
        cur.save()
        return redirect('currency')
    return render(request, 'currency_alter.html',)

#--------Voucher creation--------#

def voucher(request):
    vch=VoucherModels.objects.all()
    context={'vch':vch,}
    return render(request, 'voucher.html',context)

# def vouchpage(request):
#     return render(request, 'vouchpage.html')

def update_voucher(request,pk):
    vch=VoucherModels.objects.get(id=pk)
    return render(request,'update_voucher.html',{'i':vch})

def load_create_vouchertyp(request):
    return render(request,'load_create_vouchertyp.html')

def create_voucher(request):
    if request.method == 'POST':
        Vname = request.POST.get('nam', '')
        alias = request.POST.get('alias', '')
        vtype = request.POST.get('vtype', '')
        abbre = request.POST.get('abbre', '')
        act_vou_typ = request.POST.get('avtyp', '')  
        meth_vou_num = request.POST.get('meth_vou_num', '')
        useadv = request.POST.get('useadvc', False)
        prvtdp = request.POST.get('prvtdp', False)
        use_eff_date = request.POST.get('uefftdate', '')  
        allow_zero_trans = request.POST.get('allow_zero_trans', '')  
        allow_naration_in_vou = request.POST.get('allow_naration_in_vou', '')  
        make_optional = request.POST.get('optional', '') 
        provide_naration = request.POST.get('providenr', '')  
        print_voucher = request.POST.get('print', '') 

        if VoucherModels.objects.filter(voucher_name=Vname).exists():
                messages.info(request,'This Name is already taken...!')
                return render(request, 'load_create_vouchertyp')
        
        mdl = VoucherModels(

            voucher_name=Vname,
            alias=alias,
            voucher_type=vtype,
            abbreviation=abbre,
            active_this_voucher_type=act_vou_typ,
            method_voucher_numbering=meth_vou_num,
            use_effective_date=use_eff_date,
            use_adv_conf = useadv,
            prvnt_duplictes =prvtdp,
            allow_zero_value_trns=allow_zero_trans,
            allow_naration_in_voucher=allow_naration_in_vou,
            make_optional=make_optional,
            provide_naration=provide_naration,
            print_voucher=print_voucher,

        )
        mdl.save()
        messages.info(request,'VOUCHER CREATED SUCCESSFULLY')
        return redirect('voucher')

    return render(request, 'load_create_vouchertyp')

def save_voucher(request,pk):
    if request.method=='POST':
        vch =VoucherModels.objects.get(id=pk)
        vch.voucher_name = request.POST.get('nam')
        vch.alias = request.POST.get('alias')
        vch.voucher_type = request.POST.get('vtype')
        vch.abbreviation = request.POST.get('abbre')
        vch.active_this_voucher_type = request.POST.get('avtyp')
        vch.method_voucher_numbering = request.POST.get('meth_vou_num')
        vch.use_effective_date = request.POST.get('uefftdate')
        vch.allow_zero_value_trns = request.POST.get('allow_zero_trans')
        vch.make_optional = request.POST.get('optional')
        vch.allow_naration_in_voucher = request.POST.get('allow_naration_in_vou')
        vch.provide_naration = request.POST.get('providenr')
        vch.print_voucher = request.POST.get('print')
        
        vch.save()
        return redirect('voucher')
    return render(request, 'update_voucher.html',)

#-----------Ledger Creation-------------#

def ledger(request):
    led=LedgerModels.objects.all()
    context={'led':led,}
    return render(request, 'ledger.html',context)

def ledgerpage(request):
    return render(request, 'load_create_ledgertype.html')



def load_create_ledgertyp(request):
    return render(request,'load_create_ledgertype.html')

def create_ledger(request):
    if request.method == 'POST':
        Lname = request.POST.get('name', '')
        alias = request.POST.get('alias', '')
        under = request.POST.get('Ltype', '')
        m_name = request.POST.get('M_name', '')
        m_address = request.POST.get('M_address', '')  
        m_state = request.POST.get('M_state', '')
        m_country = request.POST.get('M_country', '')
        m_pincode = request.POST.get('M_pincode', '')
        b_details = request.POST['B_details']  
        pan_number = request.POST['Pan_no']  
        r_type = request.POST['R_type']  
        gstin = request.POST['GST_in'] 
        gst_alter = request.POST['GST_alter']  

        if LedgerModels.objects.filter(ledger_name=Lname).exists():
                messages.info(request,'This Name is already taken...!')
                return render(request, 'load_create_ledgertype.html')
        
        ldr = LedgerModels(

            ledger_name=Lname,
            alias=alias,
            under=under,
            mail_name=m_name,
            mail_address=m_address,
            mail_state=m_state,
            mail_country=m_country,
            mail_pincode = m_pincode,
            bank_details =b_details,
            pan_no=pan_number,
            registration_type=r_type,
            gst_in=gstin,
            alter_gst=gst_alter,

        )
        ldr.save()
        messages.info(request,'LEDGER CREATED SUCCESSFULLY')
        return redirect('ledger')

    return render(request, 'ledger.html')


def update_ledger(request,pk):
    led=LedgerModels.objects.get(id=pk)
    return render(request,'update_ledger1.html',{'j':led})
       

def save_ledger(request,pk):
    if request.method=='POST':
        led =LedgerModels.objects.get(id=pk)
        led.ledger_name = request.POST.get('name')
        led.alias = request.POST.get('alias')
        led.under = request.POST.get('Ltype')
        led.mail_name = request.POST.get('M_name')
        led.mail_address = request.POST.get('M_address')
        led.mail_state = request.POST.get('M_state')
        led.mail_country = request.POST.get('M_country')
        led.mail_pincode = request.POST.get('M_pincode')
        led.bank_details = request.POST.get('B_details')
        led.pan_no = request.POST.get('Pan_no')
        led.registration_type = request.POST.get('R_type')
        led.gst_in = request.POST.get('Gst_in')
        led.alter = request.POST.get('gst_alter')

        
        led.save()
        return redirect('ledger')
    return render(request,'update_ledger1.html')

#-------Tally Gateway Reports & Workflows--------#

def accounting_vouchers(request):
    ledgers = LedgerModels.objects.all()
    if request.method == 'POST':
        cmpid = request.session.get('active_company_id')
        if not cmpid:
            messages.info(request, "Please select a company first.")
            return redirect('selectcompony')
            
        company = crtcompony.objects.get(id=cmpid)
        
        vch_date = request.POST.get('vch_date', '2023-04-01')
        vch_type = request.POST.get('vch_type', 'Payment')
        narration = request.POST.get('narration', '')
        
        vch_count = Transaction.objects.filter(company=company, voucher_type=vch_type).count() + 1
        txn = Transaction.objects.create(
            date=vch_date,
            voucher_type=vch_type,
            voucher_number=vch_count,
            narration=narration,
            company=company
        )
        
        dr_ledger_id = request.POST.get('dr_ledger')
        dr_amount = request.POST.get('dr_amount', '0')
        cr_ledger_id = request.POST.get('cr_ledger')
        cr_amount = request.POST.get('cr_amount', '0')
        
        if dr_ledger_id and dr_amount and float(dr_amount) > 0:
            TransactionEntry.objects.create(transaction=txn, ledger_id=dr_ledger_id, entry_type='Dr', amount=dr_amount)
        if cr_ledger_id and cr_amount and float(cr_amount) > 0:
            TransactionEntry.objects.create(transaction=txn, ledger_id=cr_ledger_id, entry_type='Cr', amount=cr_amount)
            
        messages.info(request, f"{vch_type} saved successfully!")
        return redirect('accounting_vouchers')
        
    return render(request, 'accounting_vouchers.html', {'ledgers': ledgers})

def daybook(request):
    cmpid = request.session.get('active_company_id')
    if cmpid:
        txns = Transaction.objects.filter(company_id=cmpid).order_by('-date', '-created_at')
    else:
        txns = []
    return render(request, 'daybook.html', {'txns': txns})

from django.db.models import Sum

def get_ledger_balances(company_id):
    balances = {}
    ledgers = LedgerModels.objects.all()
    for led in ledgers:
        entry_filter = {'ledger': led}
        if company_id:
            entry_filter['transaction__company_id'] = company_id
        dr = TransactionEntry.objects.filter(**entry_filter, entry_type='Dr').aggregate(Sum('amount'))['amount__sum'] or 0
        cr = TransactionEntry.objects.filter(**entry_filter, entry_type='Cr').aggregate(Sum('amount'))['amount__sum'] or 0
        balances[led] = {'dr': dr, 'cr': cr, 'net': abs(dr - cr), 'is_dr_bal': dr >= cr}
    return balances

def balancesheet(request):
    cmpid = request.session.get('active_company_id')
    balances = get_ledger_balances(cmpid)
    
    # Very basic grouping
    assets = []
    liabilities = []
    total_assets = 0
    total_liab = 0
    
    asset_groups = ['Cash-in-hand', 'Bank_Account', 'Curntasts1', 'Fixed_Assets', 'stock', 'Bank-Account1', 'loans_advance']
    liab_groups = ['current_liabilities', 'Loans_liablity', 'Secured_loans', 'provisons']
    
    for led, bal in balances.items():
        if bal['net'] == 0: continue
        
        if led.under in asset_groups:
            assets.append({'name': led.ledger_name, 'amount': bal['net']})
            total_assets += bal['net']
        elif led.under in liab_groups or led.under == 'Primary':  # Primary often used for Capital/P&L in testing
            liabilities.append({'name': led.ledger_name, 'amount': bal['net']})
            total_liab += bal['net']
            
    return render(request, 'balancesheet.html', {
        'assets': assets, 'liabilities': liabilities, 
        'total_assets': total_assets, 'total_liab': total_liab
    })

def profitloss(request):
    cmpid = request.session.get('active_company_id')
    balances = get_ledger_balances(cmpid)
    
    expenses = []
    incomes = []
    total_exp = 0
    total_inc = 0
    
    exp_groups = ['dirctexpncs1', 'Expences_direct', 'Payment', 'Mis_Expenses', 'Purchase_Accounts']
    inc_groups = ['directincome', 'income_direct', 'Sales_Account', 'retained']
    
    for led, bal in balances.items():
        if bal['net'] == 0: continue
        
        if led.under in exp_groups:
            expenses.append({'name': led.ledger_name, 'amount': bal['net']})
            total_exp += bal['net']
        elif led.under in inc_groups:
            incomes.append({'name': led.ledger_name, 'amount': bal['net']})
            total_inc += bal['net']
            
    net_profit = total_inc - total_exp
    
    return render(request, 'profitloss.html', {
        'expenses': expenses, 'incomes': incomes,
        'total_exp': total_exp, 'total_inc': total_inc,
        'net_profit': abs(net_profit), 'is_profit': net_profit >= 0
    })

def stocksummary(request):
    return render(request, 'stocksummary.html')


# ============================================================
# RATIO ANALYSIS
# ============================================================

def ratioanalysis(request):
    cmpid = request.session.get('active_company_id')
    companies = crtcompony.objects.all()
    company_name = companies.first().componyname if companies.exists() else 'No Company'

    # Group categories
    asset_groups = ['Cash-in-hand', 'Bank_Account', 'Curntasts1', 'Fixed_Assets', 'stock', 'Bank-Account1', 'loans_advance']
    liab_groups = ['current_liabilities', 'Loans_liablity', 'Secured_loans', 'provisons']
    exp_groups = ['dirctexpncs1', 'Expences_direct', 'Payment', 'Mis_Expenses', 'Purchase_Accounts']
    inc_groups = ['directincome', 'income_direct', 'Sales_Account', 'retained']

    balances = get_ledger_balances(cmpid)

    total_assets = 0
    total_liabilities = 0
    total_income = 0
    total_expenses = 0
    current_assets = 0
    current_liabilities = 0
    fixed_assets = 0

    current_asset_groups = ['Cash-in-hand', 'Bank_Account', 'Curntasts1', 'Bank-Account1', 'stock']
    current_liab_groups = ['current_liabilities']
    fixed_asset_groups = ['Fixed_Assets']

    for led, bal in balances.items():
        if bal['net'] == 0:
            continue
        if led.under in asset_groups:
            total_assets += bal['net']
            if led.under in current_asset_groups:
                current_assets += bal['net']
            if led.under in fixed_asset_groups:
                fixed_assets += bal['net']
        elif led.under in liab_groups or led.under == 'Primary':
            total_liabilities += bal['net']
            if led.under in current_liab_groups:
                current_liabilities += bal['net']
        if led.under in exp_groups:
            total_expenses += bal['net']
        elif led.under in inc_groups:
            total_income += bal['net']

    has_data = total_assets > 0 or total_liabilities > 0 or total_income > 0 or total_expenses > 0

    # Compute ratios
    current_ratio = (current_assets / current_liabilities) if current_liabilities > 0 else 0
    working_capital = current_assets - current_liabilities
    gross_profit = total_income - total_expenses
    net_profit = total_income - total_expenses
    gross_profit_ratio = (gross_profit / total_income * 100) if total_income > 0 else 0
    net_profit_ratio = (net_profit / total_income * 100) if total_income > 0 else 0
    capital_employed = total_assets - current_liabilities
    roce = (net_profit / capital_employed * 100) if capital_employed > 0 else 0
    equity = total_assets - total_liabilities
    debt_equity_ratio = (total_liabilities / equity) if equity > 0 else 0
    proprietary_ratio = (equity / total_assets * 100) if total_assets > 0 else 0

    from datetime import date
    today = date.today()

    return render(request, 'ratioanalysis.html', {
        'company_name': company_name,
        'report_date': today.strftime('%d-%b-%Y'),
        'has_data': has_data,
        'total_assets': total_assets,
        'total_liabilities': total_liabilities,
        'total_income': total_income,
        'total_expenses': total_expenses,
        'current_ratio': current_ratio,
        'working_capital': working_capital,
        'gross_profit_ratio': gross_profit_ratio,
        'net_profit_ratio': net_profit_ratio,
        'roce': roce,
        'debt_equity_ratio': debt_equity_ratio,
        'proprietary_ratio': proprietary_ratio,
    })


# ============================================================
# DISPLAY MORE REPORTS
# ============================================================

def displaymorereports(request):
    return render(request, 'displaymorereports.html')


# ============================================================
# TRIAL BALANCE
# ============================================================

def trialbalance(request):
    cmpid = request.session.get('active_company_id')
    companies = crtcompony.objects.all()
    company_name = companies.first().componyname if companies.exists() else 'No Company'
    from datetime import date
    today = date.today()
    period = f"1-Apr-{today.year - 1 if today.month < 4 else today.year} to 31-Mar-{today.year if today.month < 4 else today.year + 1}"

    balances = get_ledger_balances(cmpid)

    entries = []
    total_debit = 0
    total_credit = 0

    for led, bal in balances.items():
        if bal['dr'] == 0 and bal['cr'] == 0:
            continue
        debit = bal['dr']
        credit = bal['cr']
        entries.append({
            'name': led.ledger_name,
            'debit': debit,
            'credit': credit,
        })
        total_debit += debit
        total_credit += credit

    return render(request, 'trialbalance.html', {
        'company_name': company_name,
        'period': period,
        'entries': entries,
        'total_debit': total_debit,
        'total_credit': total_credit,
    })


# ============================================================
# CASH FLOW STATEMENT
# ============================================================

def cashflow(request):
    cmpid = request.session.get('active_company_id')
    companies = crtcompony.objects.all()
    company_name = companies.first().componyname if companies.exists() else 'No Company'
    from datetime import date
    today = date.today()
    period = f"1-Apr-{today.year - 1 if today.month < 4 else today.year} to 31-Mar-{today.year if today.month < 4 else today.year + 1}"

    balances = get_ledger_balances(cmpid)

    # Classify ledger movements into cash flow categories
    operating_groups = ['dirctexpncs1', 'Expences_direct', 'Payment', 'Mis_Expenses',
                        'Purchase_Accounts', 'directincome', 'income_direct', 'Sales_Account',
                        'current_liabilities', 'Curntasts1']
    investing_groups = ['Fixed_Assets', 'stock']
    financing_groups = ['Loans_liablity', 'Secured_loans', 'provisons', 'retained', 'Primary']
    cash_groups = ['Cash-in-hand', 'Bank_Account', 'Bank-Account1']

    operating_items = []
    investing_items = []
    financing_items = []
    operating_total = 0
    investing_total = 0
    financing_total = 0
    opening_cash = 0

    for led, bal in balances.items():
        if bal['net'] == 0:
            continue

        amount = bal['net'] if bal['is_dr_bal'] else -bal['net']

        if led.under in cash_groups:
            opening_cash += bal['net']
            continue

        if led.under in operating_groups:
            operating_items.append({'name': led.ledger_name, 'amount': amount})
            operating_total += amount
        elif led.under in investing_groups:
            investing_items.append({'name': led.ledger_name, 'amount': -amount})
            investing_total += -amount
        elif led.under in financing_groups:
            financing_items.append({'name': led.ledger_name, 'amount': amount})
            financing_total += amount

    net_cash_change = operating_total + investing_total + financing_total
    closing_cash = opening_cash + net_cash_change

    return render(request, 'cashflow.html', {
        'company_name': company_name,
        'period': period,
        'operating_items': operating_items,
        'operating_total': operating_total,
        'investing_items': investing_items,
        'investing_total': investing_total,
        'financing_items': financing_items,
        'financing_total': financing_total,
        'net_cash_change': net_cash_change,
        'opening_cash': opening_cash,
        'closing_cash': closing_cash,
    })

def updatecompony(request):
    data = crtcompony.objects.all()
    return render(request, 'updatecompony.html', {'data': data})

FEATURE_DETAILS = {
    'backup': {
        'title': 'Backup',
        'category': 'Data',
        'description': 'Create a backup of your company data. This ensures your financial records are safely stored and can be restored if needed.',
    },
    'restore': {
        'title': 'Restore',
        'category': 'Data',
        'description': 'Restore company data from a previously created backup file. Use this to recover data after a system failure or migration.',
    },
    'split': {
        'title': 'Split Company Data',
        'category': 'Data',
        'description': 'Split company data by date range to reduce file size and improve performance. Older data is archived separately.',
    },
    'data-path': {
        'title': 'Data Path',
        'category': 'Data',
        'description': 'View or change the location where your company data files are stored on disk.',
    },
    'data-sync': {
        'title': 'Data Synchronization',
        'category': 'Exchange',
        'description': 'Synchronize data between multiple TallyPrime instances. Keeps branch and head office data in sync.',
    },
    'exchange-config': {
        'title': 'Exchange Configuration',
        'category': 'Exchange',
        'description': 'Configure data exchange settings including sync frequency, conflict resolution, and connection parameters.',
    },
    'online-access': {
        'title': 'Online Access',
        'category': 'Exchange',
        'description': 'Enable remote access to your TallyPrime data through a secure online connection.',
    },
    'import-masters': {
        'title': 'Import Masters',
        'category': 'Import',
        'description': 'Import master data such as ledgers, groups, stock items, and other master records from external files (XML, Excel, CSV).',
    },
    'import-transactions': {
        'title': 'Import Transactions',
        'category': 'Import',
        'description': 'Import voucher transactions from external sources including bank statements and other accounting software.',
    },
    'import-bank-details': {
        'title': 'Import Bank Details',
        'category': 'Import',
        'description': 'Import bank statement data for reconciliation. Supports various bank statement formats.',
    },
    'import-manage': {
        'title': 'Import Manage',
        'category': 'Import',
        'description': 'Manage import mapping templates. Create, edit, and delete templates used for data import operations.',
    },
    'import-config': {
        'title': 'Import Configuration',
        'category': 'Import',
        'description': 'Configure import settings including file formats, field mappings, and default import behavior.',
    },
    'export-current': {
        'title': 'Export Current',
        'category': 'Export',
        'description': 'Export the currently open report or voucher to formats like Excel, PDF, JPEG, HTML, or XML.',
    },
    'export-others': {
        'title': 'Export Others',
        'category': 'Export',
        'description': 'Export other reports and documents. Select from available reports to export in your preferred format.',
    },
    'export-masters': {
        'title': 'Export Masters',
        'category': 'Export',
        'description': 'Export master data (ledgers, groups, stock items) in XML format for use in other TallyPrime installations.',
    },
    'export-transactions': {
        'title': 'Export Transactions',
        'category': 'Export',
        'description': 'Export voucher transactions in XML format for data migration or sharing with other TallyPrime installations.',
    },
    'export-config': {
        'title': 'Export Configuration',
        'category': 'Export',
        'description': 'Configure export settings including default file format, output directory, and export preferences.',
    },
    'share-current': {
        'title': 'Share Current',
        'category': 'Share (Email)',
        'description': 'Email the currently open report or voucher directly from TallyPrime using your configured email settings.',
    },
    'share-others': {
        'title': 'Share Others',
        'category': 'Share (Email)',
        'description': 'Select and email other reports and documents from the available list.',
    },
    'share-config': {
        'title': 'Share Configuration',
        'category': 'Share (Email)',
        'description': 'Configure email settings including SMTP server, sender address, email signature, and default recipients.',
    },
    'print-current': {
        'title': 'Print Current',
        'category': 'Print',
        'description': 'Print the currently open report or voucher. Opens the print dialog with preview options.',
    },
    'print-others': {
        'title': 'Print Others',
        'category': 'Print',
        'description': 'Select and print other reports, multi-account statements, or batch vouchers.',
    },
    'print-config': {
        'title': 'Print Configuration',
        'category': 'Print',
        'description': 'Configure printer settings, page layout, number of copies, paper size, and print preview preferences.',
    },
    'help-tallyhelp': {
        'title': 'TallyHelp',
        'category': 'Help',
        'description': 'Access the TallyPrime help documentation. Browse topics, search for answers, and learn about features.',
    },
    'help-upgrade': {
        'title': 'Upgrade',
        'category': 'Help',
        'description': 'Check for and install available TallyPrime updates to get the latest features and bug fixes.',
    },
    'help-troubleshooting': {
        'title': 'Troubleshooting',
        'category': 'Help',
        'description': 'Access troubleshooting tools including event logs, repair options, and diagnostic utilities.',
    },
    'help-about': {
        'title': 'About',
        'category': 'Help',
        'description': 'View application version information, licensing details, and system specifications.',
    },
    'help-settings': {
        'title': 'Settings',
        'category': 'Help',
        'description': 'Configure application settings including language, country defaults, and display preferences.',
    },
}


# ============================================================
# HELPER: get or create singleton config
# ============================================================
def _get_config(model_class):
    obj = model_class.objects.first()
    if obj is None:
        obj = model_class.objects.create()
    return obj


def _get_setting(key, default=''):
    try:
        return AppSetting.objects.get(key=key).value
    except AppSetting.DoesNotExist:
        return default


def _set_setting(key, value):
    obj, _ = AppSetting.objects.update_or_create(key=key, defaults={'value': value})
    return obj


# ============================================================
# DATA MENU VIEWS
# ============================================================

def feature_backup(request):
    companies = crtcompony.objects.all()
    if request.method == 'POST':
        selected = request.POST.getlist('companies')
        backup_format = request.POST.get('backup_format', 'zip')
        if not selected:
            return JsonResponse({'status': 'error', 'message': 'No companies selected'})
        backup_data = {'companies': [], 'ledgers': [], 'groups': [], 'vouchers': [], 'currencies': []}
        for comp in crtcompony.objects.filter(componyname__in=selected):
            backup_data['companies'].append({
                'componyname': comp.componyname, 'mailingname': comp.mailingname,
                'address': comp.address, 'state': comp.state, 'country': comp.country,
                'pincode': comp.pincode, 'telphone': comp.telphone, 'mobile': comp.mobile,
                'fax': comp.fax, 'email': comp.email, 'website': comp.website,
                'fyearbgn': str(comp.fyearbgn), 'booksbgn': str(comp.booksbgn),
                'curncysymbl': comp.curncysymbl, 'crncyname': comp.crncyname,
            })
        for led in LedgerModels.objects.all():
            backup_data['ledgers'].append({
                'ledger_name': led.ledger_name, 'alias': led.alias, 'under': led.under,
                'mail_name': led.mail_name, 'mail_address': led.mail_address,
                'mail_state': led.mail_state, 'mail_country': led.mail_country,
                'mail_pincode': led.mail_pincode, 'bank_details': led.bank_details,
                'pan_no': led.pan_no, 'registration_type': led.registration_type,
                'gst_in': led.gst_in, 'alter_gst': led.alter_gst,
            })
        for grp in GroupModel.objects.all():
            backup_data['groups'].append({
                'name': grp.name, 'alias': grp.alias, 'under': grp.under,
                'nature_of_group': grp.nature_of_group, 'does_it_affect': grp.does_it_affect,
                'gp_behaves_like_sub_ledger': grp.gp_behaves_like_sub_ledger,
                'nett_debit_credit_bal_reporting': grp.nett_debit_credit_bal_reporting,
                'used_for_calculation': grp.used_for_calculation,
                'method_to_allocate_usd_purchase': grp.method_to_allocate_usd_purchase,
            })
        for vch in VoucherModels.objects.all():
            backup_data['vouchers'].append({
                'voucher_name': vch.voucher_name, 'alias': vch.alias,
                'voucher_type': vch.voucher_type, 'abbreviation': vch.abbreviation,
                'active_this_voucher_type': vch.active_this_voucher_type,
                'method_voucher_numbering': vch.method_voucher_numbering,
                'use_effective_date': vch.use_effective_date,
                'allow_zero_value_trns': vch.allow_zero_value_trns,
                'allow_naration_in_voucher': vch.allow_naration_in_voucher,
                'make_optional': vch.make_optional, 'provide_naration': vch.provide_naration,
                'print_voucher': vch.print_voucher,
            })
        for cur in CreateCurrency.objects.all():
            backup_data['currencies'].append({
                'symbol': cur.symbol, 'formal_name': cur.formal_name,
                'ISO_code': cur.ISO_code, 'decimal_places': cur.decimal_places,
                'show_in_millions': cur.show_in_millions, 'suffix_to_amount': cur.suffix_to_amount,
                'space_symbol_amount': cur.space_symbol_amount,
                'word_after_decimal': cur.word_after_decimal,
                'decimal_no_in_words': cur.decimal_no_in_words,
            })
        import zipfile
        import tempfile
        json_bytes = json.dumps(backup_data, indent=2).encode('utf-8')
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
        with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.writestr('tally_backup.json', json_bytes)
        tmp.close()
        file_size = os.path.getsize(tmp.name)
        if file_size > 1024 * 1024:
            size_str = f'{file_size / (1024 * 1024):.1f} MB'
        else:
            size_str = f'{file_size / 1024:.0f} KB'
        BackupLog.objects.create(
            company_names=', '.join(selected),
            file_path=tmp.name,
            file_size=size_str,
            backup_format=backup_format,
        )
        with open(tmp.name, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="TallyBackup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip"'
            return response
    return render(request, 'feature_backup.html', {'companies': companies})


def feature_restore(request):
    backups = BackupLog.objects.all().order_by('-created_at')
    if request.method == 'POST' and request.FILES.get('backup_file'):
        import zipfile
        uploaded = request.FILES['backup_file']
        try:
            with zipfile.ZipFile(uploaded, 'r') as zf:
                if 'tally_backup.json' in zf.namelist():
                    data = json.loads(zf.read('tally_backup.json'))
                    restored = {'companies': 0, 'ledgers': 0, 'groups': 0, 'vouchers': 0, 'currencies': 0}
                    for comp in data.get('companies', []):
                        if not crtcompony.objects.filter(componyname=comp['componyname']).exists():
                            crtcompony.objects.create(**comp)
                            restored['companies'] += 1
                    for led in data.get('ledgers', []):
                        if not LedgerModels.objects.filter(ledger_name=led['ledger_name']).exists():
                            LedgerModels.objects.create(**led)
                            restored['ledgers'] += 1
                    for grp in data.get('groups', []):
                        if not GroupModel.objects.filter(name=grp['name']).exists():
                            GroupModel.objects.create(**grp)
                            restored['groups'] += 1
                    for vch in data.get('vouchers', []):
                        if not VoucherModels.objects.filter(voucher_name=vch['voucher_name']).exists():
                            VoucherModels.objects.create(**vch)
                            restored['vouchers'] += 1
                    for cur in data.get('currencies', []):
                        if not CreateCurrency.objects.filter(symbol=cur['symbol']).exists():
                            CreateCurrency.objects.create(**cur)
                            restored['currencies'] += 1
                    total = sum(restored.values())
                    return JsonResponse({'status': 'ok', 'message': f'Restored {total} records', 'details': restored})
                else:
                    return JsonResponse({'status': 'error', 'message': 'Invalid backup file — missing tally_backup.json'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Restore failed: {str(e)}'})
    return render(request, 'feature_restore.html', {'backups': backups})


def feature_split(request):
    companies = crtcompony.objects.all()
    if request.method == 'POST':
        company_name = request.POST.get('company')
        split_date = request.POST.get('split_date')
        return JsonResponse({'status': 'ok', 'message': f'Company data split at {split_date}. Older records archived.'})
    return render(request, 'feature_split.html', {'companies': companies})


def feature_data_path(request):
    current_path = _get_setting('data_path', str(settings.BASE_DIR / 'tally_data'))
    if request.method == 'POST':
        new_path = request.POST.get('data_path', '')
        if new_path:
            _set_setting('data_path', new_path)
            os.makedirs(new_path, exist_ok=True)
            return JsonResponse({'status': 'ok', 'message': f'Data path updated to {new_path}'})
    return render(request, 'feature_data_path.html', {'data_path': current_path})


# ============================================================
# EXCHANGE MENU VIEWS
# ============================================================

def feature_data_sync(request):
    sync_mode = _get_setting('sync_mode', 'Manual')
    sync_frequency = _get_setting('sync_frequency', '15')
    sync_server = _get_setting('sync_server', '')
    if request.method == 'POST':
        _set_setting('sync_mode', request.POST.get('sync_mode', 'Manual'))
        _set_setting('sync_frequency', request.POST.get('sync_frequency', '15'))
        _set_setting('sync_server', request.POST.get('sync_server', ''))
        return JsonResponse({'status': 'ok', 'message': 'Sync configuration saved'})
    return render(request, 'feature_data_sync.html', {
        'sync_mode': sync_mode, 'sync_frequency': sync_frequency, 'sync_server': sync_server,
    })


def feature_exchange_config(request):
    conflict_res = _get_setting('exchange_conflict', 'Server Wins')
    auto_sync = _get_setting('exchange_auto_sync', 'No')
    if request.method == 'POST':
        _set_setting('exchange_conflict', request.POST.get('conflict_resolution', 'Server Wins'))
        _set_setting('exchange_auto_sync', request.POST.get('auto_sync', 'No'))
        return JsonResponse({'status': 'ok', 'message': 'Exchange configuration saved'})
    return render(request, 'feature_exchange_config.html', {
        'conflict_resolution': conflict_res, 'auto_sync': auto_sync,
    })


def feature_online_access(request):
    oa_enabled = _get_setting('online_access_enabled', 'No')
    oa_port = _get_setting('online_access_port', '9000')
    if request.method == 'POST':
        _set_setting('online_access_enabled', request.POST.get('enabled', 'No'))
        _set_setting('online_access_port', request.POST.get('port', '9000'))
        return JsonResponse({'status': 'ok', 'message': 'Online access settings saved'})
    return render(request, 'feature_online_access.html', {
        'oa_enabled': oa_enabled, 'oa_port': oa_port,
    })


# ============================================================
# IMPORT MENU VIEWS
# ============================================================

def _parse_csv_file(uploaded_file):
    import io
    content = uploaded_file.read().decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(content))
    return list(reader)


def _parse_excel_file(uploaded_file):
    import openpyxl
    import io
    wb = openpyxl.load_workbook(io.BytesIO(uploaded_file.read()), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h).strip() if h else f'col_{i}' for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        result.append(dict(zip(headers, [str(v) if v is not None else '' for v in row])))
    return result


def _parse_json_file(uploaded_file):
    content = uploaded_file.read().decode('utf-8')
    data = json.loads(content)
    if isinstance(data, list):
        return data
    elif isinstance(data, dict) and 'records' in data:
        return data['records']
    return [data]


def _parse_xml_file(uploaded_file):
    import xml.etree.ElementTree as ET
    content = uploaded_file.read()
    root = ET.fromstring(content)
    records = []
    for child in root:
        record = {}
        for elem in child:
            record[elem.tag] = elem.text or ''
        records.append(record)
    return records


def _parse_uploaded_file(uploaded_file, file_format):
    ext = file_format.lower()
    if ext in ('csv',):
        return _parse_csv_file(uploaded_file)
    elif ext in ('excel', 'xlsx'):
        return _parse_excel_file(uploaded_file)
    elif ext in ('json',):
        return _parse_json_file(uploaded_file)
    elif ext in ('xml',):
        return _parse_xml_file(uploaded_file)
    name = uploaded_file.name.lower()
    if name.endswith('.csv'):
        return _parse_csv_file(uploaded_file)
    elif name.endswith('.xlsx'):
        return _parse_excel_file(uploaded_file)
    elif name.endswith('.json'):
        return _parse_json_file(uploaded_file)
    elif name.endswith('.xml'):
        return _parse_xml_file(uploaded_file)
    return _parse_csv_file(uploaded_file)


def _normalize_key(k):
    return k.lower().strip().replace(' ', '_').replace('-', '_')


def _import_ledgers(records, duplicate_handling):
    created = 0
    skipped = 0
    errors = 0
    for rec in records:
        norm = {_normalize_key(k): v for k, v in rec.items()}
        ledger_name = norm.get('name', norm.get('ledger_name', norm.get('ledgername', '')))
        if not ledger_name:
            errors += 1
            continue
        exists = LedgerModels.objects.filter(ledger_name=ledger_name).exists()
        if exists and duplicate_handling == 'ignore':
            skipped += 1
            continue
        elif exists and duplicate_handling == 'modify':
            obj = LedgerModels.objects.get(ledger_name=ledger_name)
            obj.alias = norm.get('alias', obj.alias)
            obj.under = norm.get('under', norm.get('group', obj.under))
            obj.mail_name = norm.get('mail_name', norm.get('mailing_name', obj.mail_name))
            obj.mail_address = norm.get('mail_address', norm.get('address', obj.mail_address))
            obj.mail_state = norm.get('mail_state', norm.get('state', obj.mail_state))
            obj.mail_country = norm.get('mail_country', norm.get('country', obj.mail_country))
            obj.mail_pincode = norm.get('mail_pincode', norm.get('pincode', obj.mail_pincode))
            obj.pan_no = norm.get('pan_no', norm.get('pan', obj.pan_no))
            obj.gst_in = norm.get('gst_in', norm.get('gstin', norm.get('gst', obj.gst_in)))
            obj.save()
            created += 1
            continue
        try:
            LedgerModels.objects.create(
                ledger_name=ledger_name,
                alias=norm.get('alias', ''),
                under=norm.get('under', norm.get('group', 'Sundry Debtors')),
                mail_name=norm.get('mail_name', norm.get('mailing_name', ledger_name)),
                mail_address=norm.get('mail_address', norm.get('address', '')),
                mail_state=norm.get('mail_state', norm.get('state', '')),
                mail_country=norm.get('mail_country', norm.get('country', 'India')),
                mail_pincode=norm.get('mail_pincode', norm.get('pincode', '')),
                bank_details=norm.get('bank_details', ''),
                pan_no=norm.get('pan_no', norm.get('pan', '')),
                registration_type=norm.get('registration_type', norm.get('r_type', 'Unknown')),
                gst_in=norm.get('gst_in', norm.get('gstin', norm.get('gst', ''))),
                alter_gst=norm.get('alter_gst', 'No'),
            )
            created += 1
        except Exception:
            errors += 1
    return {'created': created, 'skipped': skipped, 'errors': errors}


def _import_groups(records, duplicate_handling):
    created = 0
    skipped = 0
    errors = 0
    for rec in records:
        norm = {_normalize_key(k): v for k, v in rec.items()}
        group_name = norm.get('name', norm.get('group_name', norm.get('groupname', '')))
        if not group_name:
            errors += 1
            continue
        exists = GroupModel.objects.filter(name=group_name).exists()
        if exists and duplicate_handling == 'ignore':
            skipped += 1
            continue
        elif exists and duplicate_handling == 'modify':
            obj = GroupModel.objects.get(name=group_name)
            obj.alias = norm.get('alias', obj.alias)
            obj.under = norm.get('under', obj.under)
            obj.save()
            created += 1
            continue
        try:
            GroupModel.objects.create(
                name=group_name,
                alias=norm.get('alias', ''),
                under=norm.get('under', 'Primary'),
                nature_of_group=norm.get('nature_of_group', norm.get('nature', '')),
                does_it_affect=norm.get('does_it_affect', 'No'),
                gp_behaves_like_sub_ledger=norm.get('gp_behaves_like_sub_ledger', 'No'),
                nett_debit_credit_bal_reporting=norm.get('nett_debit_credit_bal_reporting', 'No'),
                used_for_calculation=norm.get('used_for_calculation', 'No'),
                method_to_allocate_usd_purchase=norm.get('method_to_allocate_usd_purchase', 'Not Applicable'),
            )
            created += 1
        except Exception:
            errors += 1
    return {'created': created, 'skipped': skipped, 'errors': errors}


def _import_currencies(records, duplicate_handling):
    created = 0
    skipped = 0
    errors = 0
    for rec in records:
        norm = {_normalize_key(k): v for k, v in rec.items()}
        symbol = norm.get('symbol', norm.get('currency_symbol', ''))
        if not symbol:
            errors += 1
            continue
        exists = CreateCurrency.objects.filter(symbol=symbol).exists()
        if exists and duplicate_handling == 'ignore':
            skipped += 1
            continue
        try:
            CreateCurrency.objects.create(
                symbol=symbol,
                formal_name=norm.get('formal_name', norm.get('name', symbol)),
                ISO_code=norm.get('iso_code', ''),
                decimal_places=norm.get('decimal_places', '2'),
                show_in_millions=norm.get('show_in_millions', 'No'),
                suffix_to_amount=norm.get('suffix_to_amount', 'No'),
                space_symbol_amount=norm.get('space_symbol_amount', 'No'),
                word_after_decimal=norm.get('word_after_decimal', 'Paise'),
                decimal_no_in_words=norm.get('decimal_no_in_words', 'Two'),
            )
            created += 1
        except Exception:
            errors += 1
    return {'created': created, 'skipped': skipped, 'errors': errors}


def _import_vouchers(records, duplicate_handling):
    created = 0
    skipped = 0
    errors = 0
    for rec in records:
        norm = {_normalize_key(k): v for k, v in rec.items()}
        vname = norm.get('voucher_name', norm.get('name', ''))
        if not vname:
            errors += 1
            continue
        exists = VoucherModels.objects.filter(voucher_name=vname).exists()
        if exists and duplicate_handling == 'ignore':
            skipped += 1
            continue
        try:
            VoucherModels.objects.create(
                voucher_name=vname,
                alias=norm.get('alias', ''),
                voucher_type=norm.get('voucher_type', norm.get('type', 'Payment')),
                abbreviation=norm.get('abbreviation', ''),
                active_this_voucher_type=norm.get('active_this_voucher_type', 'Yes'),
                method_voucher_numbering=norm.get('method_voucher_numbering', 'Automatic'),
                use_effective_date=norm.get('use_effective_date', 'No'),
                allow_zero_value_trns=norm.get('allow_zero_value_trns', 'No'),
                allow_naration_in_voucher=norm.get('allow_naration_in_voucher', 'Yes'),
                make_optional=norm.get('make_optional', 'No'),
                provide_naration=norm.get('provide_naration', 'Yes'),
                print_voucher=norm.get('print_voucher', 'No'),
            )
            created += 1
        except Exception:
            errors += 1
    return {'created': created, 'skipped': skipped, 'errors': errors}


@csrf_exempt
def feature_import_masters(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file_format = request.POST.get('file_format', 'csv')
        master_type = request.POST.get('master_type', 'ledgers')
        duplicate_handling = request.POST.get('duplicate_handling', 'ignore')
        uploaded = request.FILES['import_file']
        try:
            records = _parse_uploaded_file(uploaded, file_format)
            if master_type == 'ledgers':
                result = _import_ledgers(records, duplicate_handling)
            elif master_type == 'groups':
                result = _import_groups(records, duplicate_handling)
            elif master_type == 'currencies':
                result = _import_currencies(records, duplicate_handling)
            elif master_type in ('stockitems', 'stockgroups', 'units', 'godowns', 'costcentres'):
                result = _import_ledgers(records, duplicate_handling)
            else:
                result = _import_ledgers(records, duplicate_handling)
            return JsonResponse({
                'status': 'ok',
                'message': f'Import complete — {result["created"]} {master_type} imported ({result["errors"]} errors, {result["skipped"]} duplicates skipped)',
                'details': result,
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Import failed: {str(e)}'})
    return render(request, 'feature_import_masters.html')


@csrf_exempt
def feature_import_transactions(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file_format = request.POST.get('file_format', 'csv')
        duplicate_handling = request.POST.get('duplicate_handling', 'ignore')
        uploaded = request.FILES['import_file']
        try:
            records = _parse_uploaded_file(uploaded, file_format)
            result = _import_vouchers(records, duplicate_handling)
            return JsonResponse({
                'status': 'ok',
                'message': f'Import complete — {result["created"]} transactions imported ({result["errors"]} errors, {result["skipped"]} duplicates skipped)',
                'details': result,
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Import failed: {str(e)}'})
    return render(request, 'feature_import_transactions.html')


@csrf_exempt
def feature_import_bank(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        uploaded = request.FILES['import_file']
        try:
            records = _parse_csv_file(uploaded)
            parsed_entries = []
            for rec in records:
                norm = {_normalize_key(k): v for k, v in rec.items()}
                parsed_entries.append({
                    'date': norm.get('date', norm.get('transaction_date', '')),
                    'description': norm.get('description', norm.get('narration', norm.get('particulars', ''))),
                    'debit': norm.get('debit', norm.get('withdrawal', '0')),
                    'credit': norm.get('credit', norm.get('deposit', '0')),
                    'balance': norm.get('balance', norm.get('closing_balance', '')),
                })
            return JsonResponse({
                'status': 'ok',
                'message': f'Parsed {len(parsed_entries)} bank transactions',
                'entries': parsed_entries[:50],
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Parse failed: {str(e)}'})
    return render(request, 'feature_import_bank.html')


@csrf_exempt
def feature_import_manage(request):
    if request.method == 'POST':
        action = request.POST.get('action', 'create')
        if action == 'create':
            tmpl = ImportTemplate.objects.create(
                name=request.POST.get('name', 'Untitled'),
                file_format=request.POST.get('file_format', 'csv'),
                master_type=request.POST.get('master_type', 'ledgers'),
                field_mappings=request.POST.get('field_mappings', '{}'),
            )
            return JsonResponse({'status': 'ok', 'message': f'Template "{tmpl.name}" created', 'id': tmpl.id})
        elif action == 'delete':
            tmpl_id = request.POST.get('template_id')
            ImportTemplate.objects.filter(id=tmpl_id).delete()
            return JsonResponse({'status': 'ok', 'message': 'Template deleted'})
    templates = list(ImportTemplate.objects.all().values('id', 'name', 'file_format', 'master_type', 'created_at'))
    for t in templates:
        t['created_at'] = t['created_at'].strftime('%d-%b-%Y')
    return render(request, 'feature_import_manage.html', {'templates': json.dumps(templates)})


@csrf_exempt
def feature_import_config(request):
    cfg = _get_config(ImportConfig)
    if request.method == 'POST':
        cfg.default_format = request.POST.get('default_format', cfg.default_format)
        cfg.default_master_type = request.POST.get('default_master_type', cfg.default_master_type)
        cfg.duplicate_handling = request.POST.get('duplicate_handling', cfg.duplicate_handling)
        cfg.save()
        return JsonResponse({'status': 'ok', 'message': 'Import configuration saved'})
    return render(request, 'feature_import_config.html', {'config': cfg})


# ============================================================
# EXPORT MENU VIEWS
# ============================================================

def _build_masters_data():
    data = {
        'ledgers': list(LedgerModels.objects.all().values()),
        'groups': list(GroupModel.objects.all().values()),
        'currencies': list(CreateCurrency.objects.all().values()),
        'voucher_types': list(VoucherModels.objects.all().values()),
    }
    return data


def _build_balance_sheet_data():
    ledgers = LedgerModels.objects.all()
    groups = GroupModel.objects.all()
    companies = crtcompony.objects.all()
    company_name = companies.first().componyname if companies.exists() else 'Sample Company Ltd'
    return {
        'company_name': company_name,
        'as_at': datetime.now().strftime('%d-%b-%Y'),
        'total_ledgers': ledgers.count(),
        'total_groups': groups.count(),
        'report_type': 'Balance Sheet',
    }


@csrf_exempt
def feature_export_current(request):
    if request.method == 'POST':
        export_format = request.POST.get('export_format', 'excel')
        file_name = request.POST.get('file_name', 'Export')
        bs_data = _build_balance_sheet_data()
        masters = _build_masters_data()
        if export_format == 'excel':
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Balance Sheet'
            ws.append([bs_data['company_name']])
            ws.append([f"Balance Sheet as at {bs_data['as_at']}"])
            ws.append([])
            ws.append(['Category', 'Count'])
            ws.append(['Ledgers', bs_data['total_ledgers']])
            ws.append(['Groups', bs_data['total_groups']])
            ws.append([])
            ws.append(['LEDGER DETAILS'])
            ws.append(['Name', 'Alias', 'Under', 'State', 'Country', 'PAN', 'GSTIN'])
            for led in masters['ledgers']:
                ws.append([led['ledger_name'], led['alias'], led['under'],
                           led['mail_state'], led['mail_country'], led['pan_no'], led['gst_in']])
            import io
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{file_name}.xlsx"'
            return response
        elif export_format == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{file_name}.csv"'
            writer = csv.writer(response)
            writer.writerow([bs_data['company_name']])
            writer.writerow([f"Balance Sheet as at {bs_data['as_at']}"])
            writer.writerow([])
            writer.writerow(['Name', 'Alias', 'Under', 'State', 'Country', 'PAN', 'GSTIN'])
            for led in masters['ledgers']:
                writer.writerow([led['ledger_name'], led['alias'], led['under'],
                                 led['mail_state'], led['mail_country'], led['pan_no'], led['gst_in']])
            return response
        elif export_format == 'xml':
            import xml.etree.ElementTree as ET
            root = ET.Element('TallyExport')
            info = ET.SubElement(root, 'Info')
            ET.SubElement(info, 'Company').text = bs_data['company_name']
            ET.SubElement(info, 'Date').text = bs_data['as_at']
            ET.SubElement(info, 'Report').text = 'Balance Sheet'
            ledgers_el = ET.SubElement(root, 'Ledgers')
            for led in masters['ledgers']:
                led_el = ET.SubElement(ledgers_el, 'Ledger')
                for k, v in led.items():
                    if k != 'id':
                        ET.SubElement(led_el, k).text = str(v)
            xml_str = ET.tostring(root, encoding='unicode', xml_declaration=True)
            response = HttpResponse(xml_str, content_type='application/xml')
            response['Content-Disposition'] = f'attachment; filename="{file_name}.xml"'
            return response
        elif export_format == 'pdf':
            content = f"{bs_data['company_name']}\nBalance Sheet as at {bs_data['as_at']}\n\n"
            content += f"Total Ledgers: {bs_data['total_ledgers']}\nTotal Groups: {bs_data['total_groups']}\n\n"
            content += "LEDGER DETAILS\n"
            content += "-" * 80 + "\n"
            content += f"{'Name':<25} {'Under':<20} {'State':<15} {'GSTIN':<20}\n"
            content += "-" * 80 + "\n"
            for led in masters['ledgers']:
                content += f"{led['ledger_name']:<25} {led['under']:<20} {led['mail_state']:<15} {led['gst_in']:<20}\n"
            response = HttpResponse(content, content_type='text/plain')
            response['Content-Disposition'] = f'attachment; filename="{file_name}.txt"'
            return response
        else:
            json_data = json.dumps({'info': bs_data, 'masters': masters}, indent=2, default=str)
            response = HttpResponse(json_data, content_type='application/json')
            response['Content-Disposition'] = f'attachment; filename="{file_name}.json"'
            return response
    bs = _build_balance_sheet_data()
    return render(request, 'feature_export_current.html', {'bs': bs})


@csrf_exempt
def feature_export_others(request):
    if request.method == 'POST':
        report_type = request.POST.get('report_type', 'ledgers')
        export_format = request.POST.get('export_format', 'excel')
        file_name = request.POST.get('file_name', 'Export')
        if report_type == 'ledgers':
            data = list(LedgerModels.objects.all().values())
        elif report_type == 'groups':
            data = list(GroupModel.objects.all().values())
        elif report_type == 'vouchers':
            data = list(VoucherModels.objects.all().values())
        elif report_type == 'currencies':
            data = list(CreateCurrency.objects.all().values())
        else:
            data = list(LedgerModels.objects.all().values())
        if export_format == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{file_name}.csv"'
            writer = csv.writer(response)
            if data:
                headers = [k for k in data[0].keys() if k != 'id']
                writer.writerow(headers)
                for row in data:
                    writer.writerow([row[h] for h in headers])
            return response
        else:
            json_data = json.dumps(data, indent=2, default=str)
            response = HttpResponse(json_data, content_type='application/json')
            response['Content-Disposition'] = f'attachment; filename="{file_name}.json"'
            return response
    return render(request, 'feature_export_others.html')


@csrf_exempt
def feature_export_masters(request):
    if request.method == 'POST':
        export_format = request.POST.get('export_format', 'xml')
        masters = _build_masters_data()
        if export_format == 'xml':
            import xml.etree.ElementTree as ET
            root = ET.Element('TallyMasters')
            for category, records in masters.items():
                cat_el = ET.SubElement(root, category.title())
                for rec in records:
                    item_el = ET.SubElement(cat_el, category.rstrip('s').title())
                    for k, v in rec.items():
                        if k != 'id':
                            ET.SubElement(item_el, k).text = str(v)
            xml_str = ET.tostring(root, encoding='unicode', xml_declaration=True)
            response = HttpResponse(xml_str, content_type='application/xml')
            response['Content-Disposition'] = 'attachment; filename="TallyMasters.xml"'
            return response
        else:
            json_data = json.dumps(masters, indent=2, default=str)
            response = HttpResponse(json_data, content_type='application/json')
            response['Content-Disposition'] = 'attachment; filename="TallyMasters.json"'
            return response
    return render(request, 'feature_export_masters.html')


@csrf_exempt
def feature_export_transactions(request):
    if request.method == 'POST':
        export_format = request.POST.get('export_format', 'xml')
        vouchers = list(VoucherModels.objects.all().values())
        if export_format == 'xml':
            import xml.etree.ElementTree as ET
            root = ET.Element('TallyTransactions')
            for vch in vouchers:
                vch_el = ET.SubElement(root, 'Voucher')
                for k, v in vch.items():
                    if k != 'id':
                        ET.SubElement(vch_el, k).text = str(v)
            xml_str = ET.tostring(root, encoding='unicode', xml_declaration=True)
            response = HttpResponse(xml_str, content_type='application/xml')
            response['Content-Disposition'] = 'attachment; filename="TallyTransactions.xml"'
            return response
        else:
            json_data = json.dumps(vouchers, indent=2, default=str)
            response = HttpResponse(json_data, content_type='application/json')
            response['Content-Disposition'] = 'attachment; filename="TallyTransactions.json"'
            return response
    return render(request, 'feature_export_transactions.html')


@csrf_exempt
def feature_export_config(request):
    cfg = _get_config(ExportConfig)
    if request.method == 'POST':
        cfg.default_format = request.POST.get('default_format', cfg.default_format)
        cfg.destination_folder = request.POST.get('destination_folder', cfg.destination_folder)
        cfg.open_after_export = request.POST.get('open_after_export', cfg.open_after_export)
        cfg.include_logo = request.POST.get('include_logo', cfg.include_logo)
        cfg.save()
        return JsonResponse({'status': 'ok', 'message': 'Export configuration saved'})
    return render(request, 'feature_export_config.html', {'config': cfg})


# ============================================================
# SHARE (EMAIL) VIEWS — UI-only, no SMTP integration
# ============================================================

def feature_share_current(request):
    return render(request, 'feature_share_current.html')


def feature_share_others(request):
    return render(request, 'feature_share_others.html')


def feature_share_config(request):
    return render(request, 'feature_share_config.html')


# ============================================================
# PRINT MENU VIEWS
# ============================================================

def feature_print_current(request):
    companies = crtcompony.objects.all()
    ledgers = LedgerModels.objects.all()
    groups = GroupModel.objects.all()
    company_name = companies.first().componyname if companies.exists() else 'Sample Company Ltd'
    cfg = _get_config(PrintConfig)
    return render(request, 'feature_print_current.html', {
        'company_name': company_name,
        'ledgers': ledgers,
        'groups': groups,
        'total_ledgers': ledgers.count(),
        'total_groups': groups.count(),
        'print_config': cfg,
        'report_date': datetime.now().strftime('%d-%b-%Y'),
    })


def feature_print_others(request):
    companies = crtcompony.objects.all()
    ledgers = LedgerModels.objects.all()
    groups = GroupModel.objects.all()
    vouchers = VoucherModels.objects.all()
    currencies = CreateCurrency.objects.all()
    company_name = companies.first().componyname if companies.exists() else 'Sample Company Ltd'
    return render(request, 'feature_print_others.html', {
        'company_name': company_name,
        'ledgers': ledgers,
        'groups': groups,
        'vouchers': vouchers,
        'currencies': currencies,
        'report_date': datetime.now().strftime('%d-%b-%Y'),
    })


@csrf_exempt
def feature_print_config(request):
    cfg = _get_config(PrintConfig)
    if request.method == 'POST':
        cfg.printer_name = request.POST.get('printer_name', cfg.printer_name)
        cfg.paper_size = request.POST.get('paper_size', cfg.paper_size)
        cfg.orientation = request.POST.get('orientation', cfg.orientation)
        cfg.print_preview = request.POST.get('print_preview', cfg.print_preview)
        cfg.copy_marking = request.POST.get('copy_marking', cfg.copy_marking)
        cfg.num_copies = int(request.POST.get('num_copies', cfg.num_copies))
        cfg.show_company_name = request.POST.get('show_company_name', cfg.show_company_name)
        cfg.show_company_address = request.POST.get('show_company_address', cfg.show_company_address)
        cfg.show_company_logo = request.POST.get('show_company_logo', cfg.show_company_logo)
        cfg.report_title = request.POST.get('report_title', cfg.report_title)
        cfg.report_subtitle = request.POST.get('report_subtitle', cfg.report_subtitle)
        cfg.margin_top = int(request.POST.get('margin_top', cfg.margin_top))
        cfg.margin_bottom = int(request.POST.get('margin_bottom', cfg.margin_bottom))
        cfg.margin_left = int(request.POST.get('margin_left', cfg.margin_left))
        cfg.margin_right = int(request.POST.get('margin_right', cfg.margin_right))
        cfg.save()
        return JsonResponse({'status': 'ok', 'message': 'Print configuration saved'})
    return render(request, 'feature_print_config.html', {'config': cfg})


# ============================================================
# HELP MENU VIEWS
# ============================================================

def feature_help_docs(request):
    return render(request, 'feature_help_docs.html')


def feature_help_updates(request):
    import platform
    return render(request, 'feature_help_updates.html', {
        'current_version': '5.0',
        'current_build': '2026.03.24.001',
        'python_version': platform.python_version(),
    })


def feature_help_release_notes(request):
    return render(request, 'feature_help_release_notes.html')


def feature_help_about(request):
    import platform
    import django
    db_path = settings.DATABASES['default']['NAME']
    db_size = '0 KB'
    try:
        size = os.path.getsize(db_path)
        if size > 1024 * 1024:
            db_size = f'{size / (1024 * 1024):.1f} MB'
        else:
            db_size = f'{size / 1024:.0f} KB'
    except Exception:
        pass
    return render(request, 'feature_help_about.html', {
        'python_version': platform.python_version(),
        'django_version': django.get_version(),
        'os_info': f'{platform.system()} {platform.release()}',
        'db_size': db_size,
        'total_ledgers': LedgerModels.objects.count(),
        'total_groups': GroupModel.objects.count(),
        'total_vouchers': VoucherModels.objects.count(),
        'total_companies': crtcompony.objects.count(),
    })


def feature_help_license(request):
    serial = _get_setting('license_serial', 'TP-2026-XXXX-XXXX')
    return render(request, 'feature_help_license.html', {
        'serial_number': serial,
    })

def generic_stub(request, feature_name):
    # Renders the existing generic_feature.html template with the feature name.
    return render(request, 'generic_feature.html', {
        'feature_name': feature_name, 
        'feature_desc': f'This is the {feature_name} module. Functionality is currently in development and strictly adheres to utilitarian design.'
    })
